from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from django.http import HttpResponse
from django.db import transaction

import openpyxl
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation
from datetime import date
import csv
import io
import re
import unicodedata

from .models import Categoria, Nutricional, Producto
from .forms import ProductoForm, CategoriaForm, ImportarInventarioEleventaForm
from django.contrib.auth.decorators import login_required, permission_required
# ----------------------------------------
# VISTAS CRUD
# ----------------------------------------

# --- CRUD de Categorías ---
@login_required
@permission_required('catalogo.view_categoria', raise_exception=True)
def categoria_list(request):
    nombre_filtro = request.GET.get('nombre', '')
    categorias = Categoria.objects.all()
    if nombre_filtro:
        categorias = categorias.filter(nombre__icontains=nombre_filtro)

    # PAGINACIÓN
    from django.core.paginator import Paginator
    paginator = Paginator(categorias, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'catalogo/categoria_list.html', {'page_obj': page_obj, 'nombre_filtro': nombre_filtro})

@login_required
@permission_required('catalogo.add_categoria', raise_exception=True)
def categoria_create(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría creada exitosamente.')
            return redirect('catalogo:categoria_list')
    else:
        form = CategoriaForm()
    return render(request, 'catalogo/categoria_form.html', {'form': form})

@login_required
@permission_required('catalogo.change_categoria', raise_exception=True)
def categoria_update(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada exitosamente.')
            return redirect('catalogo:categoria_list')
    else:
        form = CategoriaForm(instance=categoria)
    return render(request, 'catalogo/categoria_form.html', {'form': form})

@login_required
@permission_required('catalogo.delete_categoria', raise_exception=True)
def categoria_delete(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        nombre_cat = categoria.nombre
        categoria.delete()
        messages.success(request, f'Categoría {nombre_cat} eliminada.')
        return redirect('catalogo:categoria_list')
    return render(request, 'catalogo/categoria_confirm_delete.html', {'object': categoria})


# --- CRUD de Productos ---
@login_required
@permission_required('catalogo.view_producto', raise_exception=True)
def producto_list(request):
    nombre_filtro = request.GET.get('nombre', '')
    productos = Producto.objects.all().select_related('Categorias')
    if nombre_filtro:
        productos = productos.filter(nombre__icontains=nombre_filtro)

    from django.core.paginator import Paginator
    paginator = Paginator(productos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'catalogo/producto_list.html', {'page_obj': page_obj, 'nombre_filtro': nombre_filtro})

@login_required
@permission_required('catalogo.view_producto', raise_exception=True)
def producto_detail(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    return render(request, 'catalogo/producto_detail.html', {'producto': producto})

@login_required
@permission_required('catalogo.add_producto', raise_exception=True)
def producto_create(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            producto = form.save(commit=False)
            producto.creado = timezone.now()
            producto.save()
            messages.success(request, 'Producto creado exitosamente.')
            return redirect('catalogo:producto_list')
    else:
        form = ProductoForm()
    return render(request, 'catalogo/producto_form.html', {'form': form})

@login_required
@permission_required('catalogo.change_producto', raise_exception=True)
def producto_update(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            producto_actualizado = form.save(commit=False)
            producto_actualizado.modificado = timezone.now()
            producto_actualizado.save()
            messages.success(request, 'Producto actualizado exitosamente.')
            return redirect('catalogo:producto_list')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'catalogo/producto_form.html', {'form': form})

@login_required
@permission_required('catalogo.delete_producto', raise_exception=True)
def producto_delete(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        nombre_prod = producto.nombre
        producto.delete()
        messages.success(request, f'Producto {nombre_prod} eliminado.')
        return redirect('catalogo:producto_list')
    return render(request, 'catalogo/producto_confirm_delete.html', {'object': producto})


def _normalizar_columna(valor):
    texto = str(valor or '').strip().lower()
    texto = ''.join(
        caracter for caracter in unicodedata.normalize('NFD', texto)
        if unicodedata.category(caracter) != 'Mn'
    )
    return re.sub(r'[^a-z0-9]+', '', texto)


def _numero_entero(valor):
    if valor is None or valor == '':
        return 0
    if isinstance(valor, (int, float, Decimal)):
        return int(valor)

    texto = str(valor).strip()
    texto = re.sub(r'[^0-9,.-]', '', texto)
    if not texto:
        return 0

    if ',' in texto and '.' in texto:
        texto = texto.replace('.', '').replace(',', '.')
    elif ',' in texto:
        texto = texto.replace(',', '.')

    try:
        return int(Decimal(texto))
    except (InvalidOperation, ValueError):
        return 0


def _valor_fila(fila, columnas, nombre):
    indice = columnas.get(nombre)
    if indice is None or indice >= len(fila):
        return None
    return fila[indice]


def _detectar_columnas(encabezados):
    alias = {
        'codigo': {'codigo', 'codigobarras', 'codigodebarras', 'clave', 'sku'},
        'descripcion': {'descripcion', 'descripciondelproducto', 'producto', 'nombre', 'articulo'},
        'costo': {'costo', 'pcosto', 'costounitario', 'preciocompra', 'ultimocosto'},
        'precio': {'precio', 'pventa', 'precioventa', 'preciodeventa', 'venta', 'preciopublico'},
        'existencia': {'existencia', 'existencias', 'stock', 'cantidad'},
        'stock_minimo': {'invminimo', 'inventariominimo', 'stockminimo', 'minimo'},
        'stock_maximo': {'invmaximo', 'inventariomaximo', 'stockmaximo', 'maximo'},
        'departamento': {'departamento', 'categoria', 'rubro'},
    }
    columnas = {}

    for indice, encabezado in enumerate(encabezados):
        nombre = _normalizar_columna(encabezado)
        for campo, opciones in alias.items():
            if nombre in opciones and campo not in columnas:
                columnas[campo] = indice

    obligatorias = {'codigo', 'descripcion', 'precio', 'existencia'}
    faltantes = sorted(obligatorias - set(columnas))
    return columnas, faltantes


def _leer_filas_inventario(archivo):
    nombre = archivo.name.lower()
    if nombre.endswith('.csv'):
        contenido = archivo.read()
        try:
            texto = contenido.decode('utf-8-sig')
        except UnicodeDecodeError:
            texto = contenido.decode('latin-1')

        muestra = texto[:2048]
        try:
            dialecto = csv.Sniffer().sniff(muestra, delimiters=',;\\t')
        except csv.Error:
            dialecto = csv.excel

        return list(csv.reader(io.StringIO(texto), dialecto))

    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


@login_required
@permission_required('catalogo.add_producto', raise_exception=True)
@permission_required('catalogo.change_producto', raise_exception=True)
def producto_import_eleventa(request):
    resumen = None
    errores = []

    if request.method == 'POST':
        form = ImportarInventarioEleventaForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data['archivo']
            try:
                filas = _leer_filas_inventario(archivo)

                if not filas:
                    raise ValueError('El archivo esta vacio.')

                columnas, faltantes = _detectar_columnas(filas[0])
                if faltantes:
                    raise ValueError(
                        'Faltan columnas obligatorias: ' + ', '.join(faltantes)
                    )

                categoria_default, _ = Categoria.objects.get_or_create(
                    nombre='Inventario eleventa',
                    defaults={'descripcion': 'Productos importados desde eleventa.'}
                )
                nutricional, _ = Nutricional.objects.get_or_create(
                    ingredientes='Producto envasado segun etiqueta del fabricante',
                    defaults={
                        'tiempo_preparacion': 'Listo para llevar',
                        'proteinas': 'Segun etiqueta',
                        'azucar': 'Segun etiqueta',
                        'gluten': 'Puede contener',
                    }
                )

                creados = 0
                actualizados = 0
                omitidos = 0
                categorias_por_nombre = {'Inventario eleventa': categoria_default}

                with transaction.atomic():
                    for numero_fila, fila in enumerate(filas[1:], start=2):
                        codigo = str(_valor_fila(fila, columnas, 'codigo') or '').strip()
                        descripcion = str(_valor_fila(fila, columnas, 'descripcion') or '').strip()

                        if not codigo and not descripcion:
                            continue

                        if not codigo or not descripcion:
                            omitidos += 1
                            errores.append(f'Fila {numero_fila}: falta codigo o descripcion.')
                            continue

                        costo = _numero_entero(_valor_fila(fila, columnas, 'costo'))
                        precio = _numero_entero(_valor_fila(fila, columnas, 'precio'))
                        existencia = _numero_entero(_valor_fila(fila, columnas, 'existencia'))
                        stock_minimo = _numero_entero(_valor_fila(fila, columnas, 'stock_minimo'))
                        stock_maximo = _numero_entero(_valor_fila(fila, columnas, 'stock_maximo'))
                        departamento = str(_valor_fila(fila, columnas, 'departamento') or '').strip()

                        categoria = categoria_default
                        if departamento:
                            categoria = categorias_por_nombre.get(departamento)
                            if categoria is None:
                                categoria, _ = Categoria.objects.get_or_create(
                                    nombre=departamento[:100],
                                    defaults={'descripcion': 'Departamento importado desde eleventa.'}
                                )
                                categorias_por_nombre[departamento] = categoria

                        producto = Producto.objects.filter(codigo=codigo).first()
                        creado = False
                        if producto is None:
                            producto = Producto(codigo=codigo)
                            producto.creado = timezone.now()
                            creado = True

                        producto.nombre = descripcion[:100]
                        producto.descripcion = descripcion[:300]
                        producto.costo = costo
                        producto.precio = precio
                        producto.stock_actual = existencia
                        producto.stock_minimo = stock_minimo
                        producto.stock_maximo = stock_maximo
                        producto.tipo = 'Inventario eleventa'
                        producto.Categorias = categoria
                        producto.Nutricional = nutricional
                        producto.caducidad = date(2099, 12, 31)
                        producto.elaboracion = date.today()
                        producto.modificado = timezone.now()
                        producto.save()

                        if creado:
                            creados += 1
                        else:
                            actualizados += 1

                resumen = {
                    'creados': creados,
                    'actualizados': actualizados,
                    'omitidos': omitidos,
                }
                messages.success(
                    request,
                    f'Importacion lista: {creados} creados, {actualizados} actualizados, {omitidos} omitidos.'
                )

            except Exception as exc:
                messages.error(request, f'No se pudo importar el inventario: {exc}')
    else:
        form = ImportarInventarioEleventaForm()

    return render(request, 'catalogo/producto_import_eleventa.html', {
        'form': form,
        'resumen': resumen,
        'errores': errores[:20],
    })


# ----------------------------------------
# EXPORTACIÓN A EXCEL
# ----------------------------------------
@login_required
@permission_required('catalogo.view_producto', raise_exception=True)
def producto_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    headers = [
        "ID", "Codigo", "Nombre", "Descripción", "Marca", "Costo", "Precio",
        "Caducidad", "Elaboración", "Tipo", "Categoría", 
        "Stock Actual", "Stock Mínimo", "Stock Máximo", 
        "Presentación", "Formato"
    ]
    ws.append(headers)

    productos = Producto.objects.all().select_related('Categorias')
    for p in productos:
        ws.append([
            p.id,
            p.codigo or "",
            p.nombre,
            p.descripcion or "",
            p.marca or "",
            p.costo or 0,
            p.precio or 0,
            p.caducidad.strftime("%Y-%m-%d") if p.caducidad else "",
            p.elaboracion.strftime("%Y-%m-%d") if p.elaboracion else "",
            p.tipo,
            p.Categorias.nombre if p.Categorias else "",
            p.stock_actual or 0,
            p.stock_minimo or 0,
            p.stock_maximo or 0,
            p.presentacion or "",
            p.formato or "",
        ])

    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename=productos.xlsx'
    wb.save(response)
    return response

@login_required
@permission_required('catalogo.view_categoria', raise_exception=True)
def categoria_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Categorías"

    headers = ["ID", "Nombre", "Descripción"]
    ws.append(headers)

    categorias = Categoria.objects.all()
    for c in categorias:
        ws.append([
            c.id,
            c.nombre,
            c.descripcion or "",
        ])

    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename=categorias.xlsx'
    wb.save(response)
    return response
