from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

# --- Importaciones ---
from .models import Pedido, DetallePedido, Cliente
from .forms import ClienteForm
from catalogo.models import Producto

# --------------------
# Vistas de la Tienda (Públicas / Clientes)
# --------------------

def ver_productos(request):
    """
    Vista para que los clientes vean los productos disponibles.
    Incluye filtros persistentes y paginación.
    """
    DEFAULT_SORT = 'alpha_asc'
    DEFAULT_PER_PAGE = 9
    PER_PAGE_OPTIONS = [5, 9, 15, 30]

    SESSION_KEY_Q = 'productos_q'
    SESSION_KEY_SORT = 'productos_sort'
    SESSION_KEY_PER_PAGE = 'productos_per_page'

    # --- Búsqueda persistente ---
    if 'q' in request.GET:
        search_query = request.GET.get('q', '')
        request.session[SESSION_KEY_Q] = search_query
    else:
        search_query = request.session.get(SESSION_KEY_Q, '')

    # --- Orden persistente ---
    if 'sort' in request.GET:
        sort_by = request.GET.get('sort', DEFAULT_SORT)
        request.session[SESSION_KEY_SORT] = sort_by
    else:
        sort_by = request.session.get(SESSION_KEY_SORT, DEFAULT_SORT)

    # --- Paginación persistente ---
    if 'per_page' in request.GET:
        try:
            per_page = int(request.GET.get('per_page', DEFAULT_PER_PAGE))
            if per_page not in PER_PAGE_OPTIONS:
                per_page = DEFAULT_PER_PAGE
        except ValueError:
            per_page = DEFAULT_PER_PAGE
        request.session[SESSION_KEY_PER_PAGE] = per_page
    else:
        per_page = request.session.get(SESSION_KEY_PER_PAGE, DEFAULT_PER_PAGE)

    page_number = request.GET.get('page', 1)

    # --- Filtro base (Solo productos con stock) ---
    productos_list = Producto.objects.filter(stock_actual__gt=0)

    # --- Filtro de búsqueda ---
    if search_query:
        productos_list = productos_list.filter(
            Q(nombre__icontains=search_query) |
            Q(descripcion__icontains=search_query)
        )

    # --- Ordenamiento ---
    if sort_by == 'precio_asc':
        productos_list = productos_list.order_by('precio')
    elif sort_by == 'precio_desc':
        productos_list = productos_list.order_by('-precio')
    elif sort_by == 'alpha_desc':
        productos_list = productos_list.order_by('-nombre')
    else:
        productos_list = productos_list.order_by('nombre')

    paginator = Paginator(productos_list, per_page)
    try:
        page_obj = paginator.get_page(page_number)
    except Exception:
        page_obj = paginator.get_page(1)

    context = {
        'page_obj': page_obj,
        'paginator': paginator,
        'current_q': search_query,
        'current_sort': sort_by,
        'current_per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
    }
    return render(request, 'pedidos/ver_productos.html', context)


def agregar_al_carrito(request, pk):
    """
    Agrega un producto al carrito almacenado en la sesión.
    """
    producto = get_object_or_404(Producto, pk=pk)
    try:
        cantidad = int(request.POST.get('cantidad', 1))
    except (TypeError, ValueError):
        cantidad = 1

    carrito = request.session.get('carrito', {})
    carrito[str(pk)] = carrito.get(str(pk), 0) + cantidad
    request.session['carrito'] = carrito

    messages.success(request, f'¡Producto "{producto.nombre}" agregado al carrito!')
    
    # --- CAMBIO CLAVE AQUÍ ---
    # request.META.get('HTTP_REFERER') obtiene la URL exacta anterior (con filtros y paginación).
    # El segundo parámetro 'pedidos:ver_productos' es un respaldo por si el navegador no envía el referer.
    return redirect(request.META.get('HTTP_REFERER', 'pedidos:ver_productos'))


def ver_carrito(request):
    """
    Muestra el contenido del carrito, calcula subtotales y total general.
    """
    carrito = request.session.get('carrito', {})
    items_carrito = []
    total_carrito = 0
    total_items = 0  # Variable para contar la cantidad total de productos

    for producto_id, cantidad in carrito.items():
        producto = get_object_or_404(Producto, pk=producto_id)
        
        # --- SEGURIDAD: Convertir a enteros para cálculo matemático ---
        precio_num = int(producto.precio or 0)
        cantidad_num = int(cantidad or 0)
        
        subtotal = precio_num * cantidad_num
        
        items_carrito.append({
            'producto': producto, 
            'cantidad': cantidad_num, 
            'subtotal': subtotal
        })
        total_carrito += subtotal
        total_items += cantidad_num # Sumar al contador de items

    return render(request, 'pedidos/ver_carrito.html', {
        'items_carrito': items_carrito,
        'total_carrito': total_carrito,
        'total_items': total_items # Enviamos el total de unidades al template
    })


@login_required
@require_POST
def realizar_pedido(request):
    """
    Procesa el carrito y crea un Pedido en la base de datos.
    """
    carrito = request.session.get('carrito', {})
    if not carrito:
        messages.warning(request, "Tu carrito está vacío.")
        return redirect('pedidos:ver_productos')

    total_pedido = 0
    items_para_pedido = []
    
    # 1. Calcular total y validar productos
    for producto_id, cantidad in carrito.items():
        producto = get_object_or_404(Producto, pk=producto_id)
        
        # Validación básica de stock antes de procesar
        if cantidad > producto.stock_actual:
            messages.error(request, f"No hay suficiente stock para {producto.nombre}.")
            return redirect('pedidos:ver_carrito')
            
        precio_num = int(producto.precio or 0)
        total_pedido += precio_num * cantidad
        items_para_pedido.append((producto, cantidad, precio_num))

    # 2. Crear el Pedido
    pedido = Pedido.objects.create(usuario=request.user, total=total_pedido)
    
    # 3. Crear Detalles y Descontar Stock
    for producto, cantidad, precio in items_para_pedido:
        DetallePedido.objects.create(
            pedido=pedido, 
            producto=producto, 
            cantidad=cantidad, 
            precio=precio
        )
        # Descontar stock
        producto.stock_actual -= cantidad
        producto.save()

    # 4. Limpiar carrito
    request.session['carrito'] = {}
    
    messages.success(request, "Pedido confirmado. El pago se realiza en efectivo.")
    return redirect('pedidos:pedido_exitoso')


def pedido_exitoso(request):
    """
    Pantalla de confirmacion del pedido en efectivo.
    """
    return render(request, 'pedidos/pedido_exitoso.html')


# -------------------------------
# CRUD de Clientes (Solo Admins)
# -------------------------------

@login_required
@permission_required('pedidos.view_cliente', raise_exception=True)
def cliente_list(request):
    clientes = Cliente.objects.all().order_by('idclientes')
    return render(request, 'pedidos/cliente_list.html', {'clientes': clientes})


@login_required
@permission_required('pedidos.add_cliente', raise_exception=True)
def cliente_create(request):
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Cliente creado exitosamente.")
            return redirect('pedidos:cliente_list')
        else:
            messages.error(request, "Hubo errores al crear el cliente. Revisa los campos.")
    else:
        form = ClienteForm()

    return render(request, 'pedidos/cliente_form.html', {'form': form})


@login_required
@permission_required('pedidos.change_cliente', raise_exception=True)
def cliente_update(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        form = ClienteForm(request.POST, instance=cliente)
        if form.is_valid():
            form.save()
            messages.success(request, "Cliente actualizado exitosamente.")
            return redirect('pedidos:cliente_list')
        else:
            messages.error(request, "Error al actualizar el cliente.")
    else:
        form = ClienteForm(instance=cliente)

    return render(request, 'pedidos/cliente_form.html', {'form': form})


@login_required
@permission_required('pedidos.delete_cliente', raise_exception=True)
def cliente_delete(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == 'POST':
        cliente.delete()
        messages.success(request, "Cliente eliminado exitosamente.")
        return redirect('pedidos:cliente_list')

    return render(request, 'pedidos/cliente_confirm_delete.html', {'object': cliente})


# -------------------------------
# CRUD de Pedidos (Solo Admins)
# -------------------------------

@login_required
@permission_required('pedidos.view_pedido', raise_exception=True)
def pedido_list(request):
    """
    Lista de pedidos con opción de búsqueda.
    """
    pedidos = Pedido.objects.all().order_by('-fecha_pedido')
    
    # Filtro simple por nombre de cliente o estado
    q = request.GET.get('q')
    if q:
        pedidos = pedidos.filter(
            Q(usuario__first_name__icontains=q) |
            Q(usuario__last_name__icontains=q) |
            Q(usuario__email__icontains=q) |
            Q(estado__icontains=q)
        )

    return render(request, 'pedidos/pedido_list.html', {'pedidos': pedidos})


@login_required
@permission_required('pedidos.view_pedido', raise_exception=True)
def exportar_pedidos_excel(request):
    pedidos = Pedido.objects.all().order_by('-fecha_pedido')

    # Crear un libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pedidos"

    # Encabezados
    columnas = ["ID", "Cliente", "Correo", "Fecha del Pedido", "Total", "Estado"]
    for col_num, titulo in enumerate(columnas, 1):
        c = ws.cell(row=1, column=col_num)
        c.value = titulo
        c.font = openpyxl.styles.Font(bold=True)

    # Filas de datos
    for row_num, pedido in enumerate(pedidos, 2):
        ws.cell(row=row_num, column=1).value = pedido.id
        ws.cell(row=row_num, column=2).value = f"{pedido.usuario.first_name} {pedido.usuario.last_name}"
        ws.cell(row=row_num, column=3).value = pedido.usuario.email
        ws.cell(row=row_num, column=4).value = pedido.fecha_pedido.strftime("%d/%m/%Y %H:%M")
        ws.cell(row=row_num, column=5).value = float(pedido.total)
        ws.cell(row=row_num, column=6).value = pedido.estado

    # Ajustar ancho automático
    for col_num in range(1, len(columnas) + 1):
        columna = get_column_letter(col_num)
        ws.column_dimensions[columna].width = 20

    # Respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="pedidos.xlsx"'
    wb.save(response)
    return response


@login_required
@permission_required('pedidos.change_pedido', raise_exception=True)
@require_POST
def pedido_update_status(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    estado = request.POST.get('estado')
    estados_validos = [choice[0] for choice in Pedido.ESTADO_CHOICES]

    if estado in estados_validos:
        pedido.estado = estado
        pedido.save()
        messages.success(request, f"El estado del pedido #{pedido.id} se actualizó a '{estado}'.")
    else:
        messages.error(request, "Estado de pedido inválido.")

    return redirect('pedidos:pedido_detail', pk=pedido.id)


@login_required
@permission_required('pedidos.view_pedido', raise_exception=True)
def pedido_detail(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    
    # Calcular subtotales para mostrarlos en el template si no están guardados
    for detalle in pedido.detalles.all():
        detalle.subtotal = detalle.cantidad * detalle.precio

    can_change_pedido = request.user.has_perm('pedidos.change_pedido')
    return render(request, 'pedidos/pedido_detail.html', {
        'pedido': pedido,
        'can_change_pedido': can_change_pedido,
    })